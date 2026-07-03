import os
from openpyxl import load_workbook
import db

class Excel:
    def __init__(self, ruta_excel, nombre_hoja = "Hoja1"):
        self.ruta_excel = ruta_excel
        self.nombre_hoja = nombre_hoja
        self.wb = load_workbook(ruta_excel)

        # Verificar si la hoja existe
        if nombre_hoja not in self.wb.sheetnames:
            print(f"Error: La hoja '{nombre_hoja}' no existe.")
            return

    def modificar_celda(self, celda, nuevo_valor):
        try:
            hoja = self.wb[self.nombre_hoja]
            hoja[celda] = nuevo_valor
            self.wb.save(self.ruta_excel)
        except Exception as e:
            print(f"Ocurrió un error al modificar la celda: {celda}, {nuevo_valor}.\nError: {e}")

    def leer_celda(self, celda):
        try:
            hoja = self.wb[self.nombre_hoja]
            return hoja[celda].value
        except Exception as e:
            print(f"Ocurrió un error al leer la celda: {e}")
            return None
        
    def cerrar(self):
        self.wb.close()

    def crear_retencion(self, diccionario_modificaciones):
        for celda, valor in diccionario_modificaciones.items():
            self.modificar_celda(celda, valor)
        print("Retención creada exitosamente.")


arch = Excel("Plantilla Retenciones osteobone.ods", "Hoja1")

#inputs
date = "02/07/2026"
enterprise_name = "Osteobone C.A."
enterprise_rif = "J-12345678-9"
recipe_number = "123456789"
control_number = "987654321"
base_imponible = "1000.00"


# process varibles
scuensial_number = db.obtener_siguiente_codigo()

split_date = date.split("/")
month = split_date[1]
year = split_date[2]

numeroComprobante = f'N. DE COMPROBANTE {year}{month}{scuensial_number}'

# Definimos un diccionario para mapear los números de mes a sus nombres en español
meses = {
    "01": "ENERO",
    "02": "FEBRERO",
    "03": "MARZO",
    "04": "ABRIL",
    "05": "MAYO",
    "06": "JUNIO",
    "07": "JULIO",
    "08": "AGOSTO",
    "09": "SEPTIEMBRE",
    "10": "OCTUBRE",
    "11": "NOVIEMBRE",
    "12": "DICIEMBRE"
}

# Variable para almacenar el número de factura
diccionario_modificaciones = {
    'B6' : numeroComprobante, #Numero de Comprobante
    'M6' : date, #Fecha de Emisión
    'M10' : year, #Periodo Fiscal (AÑO)
    'N10' : meses[month], #Periodo Fiscal (MES)
    'A17' : enterprise_name, #Razon Social
    'K17' : enterprise_rif, #RIF
    'B21' : date, #Fecha de Documento
    'C21' : recipe_number, #Numero de Factura
    'D21' : control_number, #Numero de Control
    'K21' : base_imponible, #Base Imponible
}

arch.crear_retencion(diccionario_modificaciones)

print("Ejecición finalizada.")